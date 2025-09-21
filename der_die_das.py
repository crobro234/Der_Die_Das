
import os
import random
import tkinter as tk
from tkinter import messagebox

# --- Data loader: auto-detect vertical(A=word,B=article) or horizontal(row-wise) ---
def load_word_article_pairs(xlsx_path="words.xlsx", orientation="auto"):
    """
    지원 레이아웃
    1) 세로(열) 배치: A열=단어, B열=관사(der/die/das)
       A1=Haus, B1=das
       A2=Mann, B2=der
       ...
    2) 가로(행) 배치: 1행=단어들, 2행=관사들
       A1=Haus, B1=Mann, C1=Frau, ...
       A2=das,  B2=der,  C2=die,  ...

    orientation="vertical" 또는 "horizontal"로 강제 지정 가능. 기본은 "auto".
    """
    if not os.path.exists(xlsx_path):
        raise FileNotFoundError(f"Excel file not found: {xlsx_path}")

    valid_articles = {"der", "die", "das"}

    def _clean(s):
        if s is None:
            return ""
        return str(s).strip()

    # 1) pandas 시도
    df = None
    try:
        import pandas as pd
        df = pd.read_excel(xlsx_path, header=None)
        df = df.where(pd.notnull(df), None)  # NaN -> None
        is_pandas = True
    except Exception:
        is_pandas = False

    # 2) pandas 실패 시 openpyxl 폴백
    if not is_pandas:
        try:
            from openpyxl import load_workbook
            wb = load_workbook(xlsx_path, read_only=True, data_only=True)
            ws = wb.active
            # 실제 값이 있는 범위를 위에서부터 훑어서 수집
            rows = []
            empty_streak = 0
            r = 1
            while empty_streak < 10:  # 빈 행 10번 연속 나오면 종료
                row_vals = [cell.value for cell in ws[r]]
                if len(row_vals) == 0:
                    empty_streak += 1
                elif all(v is None for v in row_vals):
                    empty_streak += 1
                else:
                    empty_streak = 0
                    rows.append(row_vals)
                r += 1

            # 들쭉날쭉한 길이 -> 패딩
            maxc = max((len(rw) for rw in rows), default=0)
            for i in range(len(rows)):
                if len(rows[i]) < maxc:
                    rows[i] += [None] * (maxc - len(rows[i]))

            class _Wrapper:
                def __init__(self, data):
                    self.data = data
                @property
                def shape(self):
                    return (len(self.data), max((len(r) for r in self.data), default=0))
                def get_col(self, c):
                    return [(row[c] if c < len(row) else None) for row in self.data]
                def get_row(self, r):
                    return list(self.data[r]) if r < len(self.data) else []

            df = _Wrapper(rows)
        except Exception as e:
            raise RuntimeError(f"Failed to read Excel with pandas or openpyxl. Details: {e}")

    # 빌더들
    def build_vertical_pairs(df_like):
        pairs = []
        try:
            # pandas
            if hasattr(df_like, "iloc"):
                col_w = list(df_like.iloc[:, 0])
                col_a = list(df_like.iloc[:, 1])
            else:
                # 폴백 래퍼
                col_w = df_like.get_col(0)
                col_a = df_like.get_col(1)
        except Exception:
            return pairs

        for w, a in zip(col_w, col_a):
            w = _clean(w)
            a = _clean(a).lower()
            if w and a in valid_articles:
                pairs.append((w, a))
        return pairs

    def build_horizontal_pairs(df_like):
        pairs = []
        try:
            # pandas
            if hasattr(df_like, "iloc"):
                if df_like.shape[0] < 2:
                    return pairs
                row_words = list(df_like.iloc[0, :])
                row_arts  = list(df_like.iloc[1, :])
            else:
                # 폴백 래퍼
                if df_like.shape[0] < 2:
                    return pairs
                row_words = df_like.get_row(0)
                row_arts  = df_like.get_row(1)
        except Exception:
            return pairs

        for w, a in zip(row_words, row_arts):
            w = _clean(w)
            a = _clean(a).lower()
            if w and a in valid_articles:
                pairs.append((w, a))
        return pairs

    vertical_pairs = build_vertical_pairs(df)
    horizontal_pairs = build_horizontal_pairs(df)

    choice = orientation.lower() if isinstance(orientation, str) else "auto"
    if choice == "vertical":
        chosen = vertical_pairs
    elif choice == "horizontal":
        chosen = horizontal_pairs
    else:
        # auto: 유효쌍이 더 많은 레이아웃 채택 (동률이면 세로 우선)
        chosen = vertical_pairs if len(vertical_pairs) >= len(horizontal_pairs) else horizontal_pairs

    if not chosen:
        raise ValueError(
            "No valid (word, article) pairs found.\n"
            "- 세로(열): A열=단어, B열=der/die/das\n"
            "- 가로(행): 1행=단어들, 2행=der/die/das"
        )
    return chosen


class DerDieDasApp(tk.Tk):
    def __init__(self, pairs):
        super().__init__()
        self.title("DER DIE DAS")
        self.geometry("720x480")
        self.resizable(False, False)

        self.pairs = pairs[:]  # list of (word, article)
        self.deck = []
        self.idx = 0
        self.waiting_for_next = False
        self.score = 0
        self.seen = 0

        # Key mappings
        self.key_to_article = {"1": "der", "2": "die", "3": "das"}

        # --- UI ---
        topbar = tk.Frame(self)
        topbar.pack(fill="x", pady=(14, 0), padx=14)

        self.header = tk.Label(topbar, text="DER DIE DAS", font=("Segoe UI", 22, "bold"))
        self.header.pack(side="left")

        # 점수 표시(상단 우측)
        self.score_label = tk.Label(
            topbar, text="Score: 0 / 0 (0%)", font=("Segoe UI", 11)
        )
        self.score_label.pack(side="right")

        self.subtitle = tk.Label(
            self,
            text="Click Start to begin. Choose by buttons or press 1=der, 2=die, 3=das.",
            font=("Segoe UI", 11)
        )
        self.subtitle.pack(pady=(6, 14))

        self.word_label = tk.Label(self, text="—", font=("Segoe UI", 44, "bold"))
        self.word_label.pack(pady=16)

        # Buttons frame
        btn_frame = tk.Frame(self)
        btn_frame.pack(pady=10)

        self.btn_start = tk.Button(btn_frame, text="Start", width=12, command=self.start_quiz)
        self.btn_start.grid(row=0, column=0, padx=8)

        self.btn_next = tk.Button(btn_frame, text="Next", width=12, command=self.next_question, state="disabled")
        self.btn_next.grid(row=0, column=1, padx=8)

        # Options frame
        opt_frame = tk.Frame(self)
        opt_frame.pack(pady=14)

        self.btn_der = tk.Button(opt_frame, text="der (1)", width=14,
                                 command=lambda: self.submit("der"), state="disabled")
        self.btn_die = tk.Button(opt_frame, text="die (2)", width=14,
                                 command=lambda: self.submit("die"), state="disabled")
        self.btn_das = tk.Button(opt_frame, text="das (3)", width=14,
                                 command=lambda: self.submit("das"), state="disabled")

        # Keep order fixed: der, die, das
        self.btn_der.grid(row=0, column=0, padx=8)
        self.btn_die.grid(row=0, column=1, padx=8)
        self.btn_das.grid(row=0, column=2, padx=8)

        # Feedback & status
        self.feedback = tk.Label(self, text="", font=("Segoe UI", 13))
        self.feedback.pack(pady=(8, 4))

        self.status = tk.Label(self, text="Ready.", font=("Segoe UI", 11))
        self.status.pack()

        # 오른쪽 아래 워터마크 (작게)
        self.watermark = tk.Label(
            self, text="made by DongJu\nver 1.10",
            font=("Segoe UI", 9), fg="#777777", justify="right"
        )
        self.watermark.place(relx=1.0, rely=1.0, x=-10, y=-10, anchor="se")

        # Key bindings
        self.bind("<Key>", self.on_key)

    # --- helpers ---
    def update_score_label(self):
        acc = 0 if self.seen == 0 else round(self.score * 100 / self.seen)
        self.score_label.config(text=f"Score: {self.score} / {self.seen} ({acc}%)")

    # --- Quiz control ---
    def start_quiz(self):
        random.shuffle(self.pairs)
        self.deck = self.pairs[:]
        self.idx = 0
        self.score = 0
        self.seen = 0
        self.waiting_for_next = False
        self.btn_start.config(state="disabled")
        self.enable_options(True)
        self.btn_next.config(state="disabled")
        self.feedback.config(text="")
        self.status.config(text="Quiz started. Good luck!")
        self.update_score_label()
        self.next_question()

    def enable_options(self, enable: bool):
        state = "normal" if enable else "disabled"
        self.btn_der.config(state=state)
        self.btn_die.config(state=state)
        self.btn_das.config(state=state)

    def show_word(self, word):
        self.word_label.config(text=word)

    def next_question(self, *_):
        if self.waiting_for_next:
            self.waiting_for_next = False

        if self.idx >= len(self.deck):
            random.shuffle(self.pairs)
            self.deck = self.pairs[:]
            self.idx = 0
            self.status.config(text="Deck completed. Shuffling and continuing...")

        self.current_word, self.current_art = self.deck[self.idx]
        self.idx += 1
        self.seen += 1

        self.show_word(self.current_word)
        self.feedback.config(text="")
        self.enable_options(True)
        self.btn_next.config(state="disabled")
        self.update_score_label()

    def submit(self, guess):
        if self.waiting_for_next:
            return

        correct = (guess == self.current_art)
        if correct:
            self.score += 1
            self.feedback.config(text="Correct!", fg="#117A65")
            self.status.config(text=f"Nice! Pressed {guess.upper()}.")
            self.update_score_label()
            self.after(150, self.next_question)
        else:
            self.feedback.config(text=f"Wrong. Correct answer: {self.current_art.upper()}", fg="#C0392B")
            self.status.config(text="Press any key or Next to continue.")
            self.enable_options(False)
            self.btn_next.config(state="normal")
            self.waiting_for_next = True
            self.update_score_label()

    # --- Keyboard handling ---
    def on_key(self, event):
        key = event.keysym
        if self.waiting_for_next:
            self.next_question()
            return
        if key in ("1", "2", "3"):
            self.submit(self.key_to_article[key])


def main():
    try:
        # 세로(열) 배치를 기본으로 사용. 필요 시 orientation="horizontal"로 강제 가능.
        pairs = load_word_article_pairs("words.xlsx", orientation="auto")
    except Exception as e:
        messagebox.showerror("Error", f"Failed to load words.xlsx\n\n{e}")
        return

    app = DerDieDasApp(pairs)
    app.mainloop()


if __name__ == "__main__":
    main()
